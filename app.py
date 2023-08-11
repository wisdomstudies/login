from flask import Flask, render_template, request, redirect, session
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "your_secret_key"  # Replace with a secret key for session management

# Load user data from Excel file
def load_user_data():
    user_data = []
    wb = load_workbook('user_data.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user_data.append({
            "Username": row[0],
            "Password": row[1],
            "Name": row[2],
            "Test1": row[3],
            "Test2": row[4],
            # Add other columns here
        })
    return user_data

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']
    user_data = load_user_data()

    for user in user_data:
        if user['Username'] == username and user['Password'] == password:
            session['userData'] = user
            return redirect('/grades')
    
    return "Invalid username or password."

@app.route('/grades')
def grades():
    if 'userData' in session:
        user = session['userData']
        return render_template('grades.html', user=user)
    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)
